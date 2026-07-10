"""Excel 解析工具实现（pandas / openpyxl，大表走 DuckDB 分块）。

要点：仅产出"数据画像"，原始整表不进 LLM（红线1）；大表超阈值切 DuckDB（第7节，留 TODO）。
"""

from __future__ import annotations

import math
from typing import Any

import pandas as pd
from packages.common.config import get_settings
from packages.common.dataset_store import load_dataframe, save_dataframe
from packages.common.logging import get_logger
from packages.governance.data_boundary import resolve_policy
from packages.governance.redaction import apply_policy

from mcp_servers.excel_parser.profile import ColumnProfile, DataProfile

_log = get_logger("mcp.excel_parser")

# 默认样本行数（属"画像"范畴，可喂 LLM；见设计文档 6.1）
_SAMPLE_ROWS = 5
# 每列展示的样本值个数
_SAMPLE_VALUES = 5


class TableTooLargeError(ValueError):
    """表行数超过处理上限（大表防护：整表读回内存前先拒绝）。"""


def parse_excel(args: dict[str, Any]) -> dict[str, Any]:
    """解析 Excel 文件，落地为数据集引用（dataset_ref），不返回整表（红线1）。

    Args:
        args: file_ref（必填）、sheet、header_row、nrows（均可选）。

    Returns:
        {dataset_ref, row_count, column_count}。
    """
    file_ref: str = args["file_ref"]
    sheet: str | int = args.get("sheet", 0)
    header_row: int = args.get("header_row", 0)
    nrows: int | None = args.get("nrows")

    # 大表防护：读整表进内存前先查行数元数据，超阈值直接拒绝（防 OOM）。
    # TODO（大表）：后续支持超阈值文件时改 DuckDB 扫描/分块，而非直接拒绝。
    _guard_row_limit(file_ref, sheet, header_row, nrows)
    df = pd.read_excel(file_ref, sheet_name=sheet, header=header_row, nrows=nrows)
    dataset_ref = save_dataframe(df)
    return {
        "dataset_ref": dataset_ref,
        "row_count": int(df.shape[0]),
        "column_count": int(df.shape[1]),
    }


def infer_schema(args: dict[str, Any]) -> DataProfile:
    """推断 schema 与统计摘要，生成数据画像（DataProfile）。

    Args:
        args: dataset_ref（必填）。

    Returns:
        DataProfile —— 喂给推理模型的唯一数据视图。
    """
    dataset_ref: str = args["dataset_ref"]
    df = load_dataframe(dataset_ref)
    columns = [_profile_column(df[col]) for col in df.columns]
    sample_rows = _json_safe_records(df.head(_SAMPLE_ROWS))
    profile = DataProfile(
        dataset_ref=dataset_ref,
        row_count=int(df.shape[0]),
        column_count=int(df.shape[1]),
        columns=columns,
        sample_rows=sample_rows,
    )
    # 第1层：按数据集安全策略脱敏采样后再返回（红线1）。
    # 默认宽松；数据集 sidecar 可收紧。仅 VALUES 列的真实单元格才进入 payload。
    return apply_policy(profile, resolve_policy(dataset_ref))


def data_preview(args: dict[str, Any]) -> dict[str, Any]:
    """返回少量样本行供用户确认（前端先展示画像再分析）。"""
    dataset_ref: str = args["dataset_ref"]
    rows: int = args.get("rows", 20)
    df = load_dataframe(dataset_ref)
    return {"rows": _json_safe_records(df.head(rows))}


# ── 内部辅助 ──

def _guard_row_limit(
    file_ref: str, sheet: str | int, header_row: int, nrows: int | None
) -> None:
    """行数上限防护：openpyxl read_only 只读工作表元数据，不解压整表数据。

    上传大小上限只约束压缩后体积，高压缩比 xlsx 解开后仍可能打爆内存，
    故在 pd.read_excel 之前按 `large_table_row_threshold` 拒绝超大表。

    Args:
        file_ref: 文件路径（仅 .xlsx/.xlsm 可查；.xls 走不了 openpyxl，跳过）。
        sheet: 工作表名或序号。
        header_row: 表头行号（0 基），行数按数据行计算。
        nrows: 调用方限定的读取行数；不超阈值则无需检查。

    Raises:
        TableTooLargeError: 数据行数超过阈值。
        ValueError: 工作表不存在。
    """
    limit = get_settings().large_table_row_threshold
    if nrows is not None and nrows <= limit:
        return
    if not file_ref.lower().endswith((".xlsx", ".xlsm")):
        return

    from openpyxl import load_workbook

    wb = load_workbook(file_ref, read_only=True)
    try:
        if isinstance(sheet, str):
            if sheet not in wb.sheetnames:
                raise ValueError(f"工作表不存在: {sheet}")
            max_row = wb[sheet].max_row
        else:
            if not 0 <= sheet < len(wb.worksheets):
                raise ValueError(f"工作表不存在: {sheet}")
            max_row = wb.worksheets[sheet].max_row
    finally:
        wb.close()

    if max_row is None:
        # 个别生成器不写 dimension 元数据；跳过检查并告警，不误伤正常文件
        _log.warning("excel.row_limit.unknown", file_ref=file_ref)
        return
    data_rows = max(0, max_row - 1 - header_row)  # 去掉表头及其上方行
    if data_rows > limit:
        raise TableTooLargeError(
            f"表行数超过处理上限（约 {data_rows} 行 > {limit} 行），"
            f"请拆分文件或缩小数据范围后重试"
        )


def _dtype_name(series: pd.Series) -> str:
    """把 pandas dtype 归一为画像用的简单类型名。"""
    if pd.api.types.is_bool_dtype(series):
        return "bool"
    if pd.api.types.is_integer_dtype(series):
        return "int"
    if pd.api.types.is_float_dtype(series):
        return "float"
    if pd.api.types.is_datetime64_any_dtype(series):
        return "datetime"
    return "str"


def _profile_column(series: pd.Series) -> ColumnProfile:
    """生成单列画像。数值列附 describe 统计摘要。"""
    dtype = _dtype_name(series)
    total = len(series)
    null_ratio = float(series.isna().mean()) if total else 0.0
    distinct_count = int(series.nunique(dropna=True))
    # 先采原始样本值；随后由 governance.redaction 按数据集策略脱敏（见 infer_schema）。
    sample_values = [
        _scalar_to_str(v) for v in series.dropna().unique()[:_SAMPLE_VALUES]
    ]

    profile = ColumnProfile(
        name=str(series.name),
        dtype=dtype,
        null_ratio=null_ratio,
        distinct_count=distinct_count,
        sample_values=sample_values,
    )
    if dtype in ("int", "float"):
        desc = series.astype("float64")
        profile.min = _safe_float(desc.min())
        profile.max = _safe_float(desc.max())
        profile.mean = _safe_float(desc.mean())
        profile.std = _safe_float(desc.std())
        profile.median = _safe_float(desc.median())
    return profile


def _safe_float(value: Any) -> float | None:
    """把统计值转为 JSON 安全的 float（NaN/inf → None）。"""
    try:
        f = float(value)
    except (TypeError, ValueError):
        return None
    if math.isnan(f) or math.isinf(f):
        return None
    return round(f, 6)


def _scalar_to_str(value: Any) -> str:
    """标量转字符串（样本值用）。"""
    return "" if value is None else str(value)


def _json_safe_records(df: pd.DataFrame) -> list[dict]:
    """DataFrame → JSON 安全的记录列表（NaN→None，时间→iso 字符串）。"""
    safe = df.copy()
    for col in safe.columns:
        if pd.api.types.is_datetime64_any_dtype(safe[col]):
            safe[col] = safe[col].astype(str)
    records = safe.to_dict(orient="records")
    out: list[dict] = []
    for rec in records:
        out.append(
            {k: (None if (isinstance(v, float) and math.isnan(v)) else v) for k, v in rec.items()}
        )
    return out
